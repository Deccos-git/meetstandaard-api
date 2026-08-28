#!/usr/bin/env python3
"""
Build a versioned meetstandaard JSON document from the source workbook.

The workbook is the authoring format (10 sheets, one per aspect of the
standard); the API serves a single nested JSON document per version. This
script is the only bridge between the two, so a regenerated version is always
byte-reproducible from its source workbook.

Usage:
    python3 tools/build-meetstandaard.py \
        --xlsx "/path/to/meetstandaard.xlsx" \
        --sector energiearmoede \
        --version 0.9 \
        --released-at 2026-08-17 \
        --out functions/data/meetstandaard-energiearmoede-0.9.json

Requires: openpyxl (pip install openpyxl). Not a runtime dependency of the
Cloud Functions — the generated JSON is committed and imported directly.
"""

import argparse
import json
import re
from collections import OrderedDict, defaultdict

import openpyxl

from xlsx_common import clean, rows, slugify, split_list, to_int, to_number

SHEETS = {
    "effecten": "1-Effecten",
    "stellingen": "2-Stellingen",
    "situatieschetsen": "3-Situatieschetsen",
    "monetarisering": "4-Monetarisering-per-niveau",
    "proxies": "5-Stakeholder-proxywaarden",
    "bronnen": "6-Bronnen",
    "audit": "7-Audit",
    "aggregatie": "8-Aggregatie",
    "parameters": "9-Parameters",
    "gevoeligheid": "10-Gevoeligheid",
}

SECTOR_LABELS = {"energiearmoede": "Energiearmoede"}


def slug_from_path(path, fallback):
    """'wiki/effecten/fysieke-gezondheid.md' -> 'fysieke-gezondheid'."""
    if path and path.endswith(".md"):
        return path.rsplit("/", 1)[-1][: -len(".md")]
    return slugify(fallback)


def score_of(niveau):
    """Numeric Likert score, or None for effects scored on another scale."""
    if niveau is None:
        return None
    if isinstance(niveau, (int, float)):
        return int(niveau)
    return int(niveau) if re.fullmatch(r"\d+", str(niveau).strip()) else None


def ja_nee(value):
    """'ja'/'nee' -> True/False, leeg -> None.

    Alleen voor kolommen waarin de auteur een feit over de stelling vastlegt.
    Een leeg antwoord is niet hetzelfde als "nee": de workbooks van de
    standaarden die uit de `effects`-collectie migreren hebben deze kolom leeg
    overal waar het oude panel hem nooit heeft weggeschreven.

    Zie docs/pitfalls.md#data-ui-default-never-persisted.
    """
    if value is None:
        return None
    return str(value).strip().lower().startswith("ja")


def build_effecten(wb, sector):
    effecten = OrderedDict()
    for row in rows(wb[SHEETS["effecten"]]):
        eid = row["effect_id"]
        slug = slug_from_path(row.get("opmerkingen"), row.get("effect"))
        effecten[eid] = OrderedDict(
            id=eid,
            slug=slug,
            # Cross-standaard key, en de MSIId waarop consumenten joinen.
            #
            # `id` (EFF-01) telt per workbook opnieuw en `slug` komt uit de
            # effectnaam, dus twee sectoren met "Betere fysieke gezondheid"
            # leveren allebei `fysieke-gezondheid` op. Dat zijn niet dezelfde
            # effecten — een effect is pas hetzelfde over sectoren heen als de
            # stellingen identiek zijn — dus de sleutel moet de sector dragen.
            #
            # De versie zit er bewust niet in: een effect houdt zijn identiteit
            # over 0.9 -> 1.0 heen, en een versiegebonden sleutel zou elke
            # publicatie er voor een consument als een nieuw effect uit laten
            # zien. De versie hoort op het record, niet in de sleutel.
            #
            # De dubbele punt omdat het resultaat een Firestore-document-id moet
            # kunnen zijn (waar "/" een padscheiding is) en een URL-segment
            # zonder escapen.
            #
            # Tot 2026-08-24 werd hij alleen bij het seeden gezet, waardoor het
            # gecommitte document en het geserveerde document verschilden. Zie
            # docs/pitfalls.md#versioning-reseeding-strands-cached-clients.
            uid=f"{sector}:{slug}",
            effect=row.get("effect"),
            categorie=row.get("categorie"),
            typeEffect=row.get("type_effect"),
            definitie=row.get("definitie"),
            bronDefinitie=row.get("bron_definitie"),
            doelgroep=row.get("doelgroep"),
            relevantieSector=row.get("relevantie_sector"),
            crossSectorBenchmarkbaar=row.get("cross_sector_benchmarkbaar"),
            herkomstBestaandeStandaard=row.get("herkomst_bestaande_standaard"),
            monetariseerbaarheid=row.get("monetariseerbaarheid"),
            document=row.get("opmerkingen"),
            stellingen=[],
            situatieschetsen=[],
            monetarisering=OrderedDict(document=None, eenheid=None, niveaus=[]),
        )

    for row in rows(wb[SHEETS["stellingen"]]):
        effecten[row["effect_id"]]["stellingen"].append(
            OrderedDict(
                nummer=to_int(row.get("stelling_nummer")),
                stelling=row.get("stelling"),
                bron=row.get("bron_stelling"),
                origineleSchaal=row.get("originele_schaal"),
                gebruikteSchaal=row.get("gebruikte_schaal"),
                richting=row.get("richting"),
                # Een lege cel is *onbekend*, nooit "nee". Deze vlag bepaalt of
                # een hoge score meer of minder van het effect betekent, dus een
                # verzonnen `false` draait stil elk antwoord op die stelling om.
                negatiefGeformuleerd=ja_nee(row.get("negatief_geformuleerd")),
                letterlijkOvergenomen=ja_nee(row.get("letterlijk_overgenomen")),
                toelichting=row.get("toelichting"),
            )
        )

    for row in rows(wb[SHEETS["situatieschetsen"]]):
        effecten[row["effect_id"]]["situatieschetsen"].append(
            OrderedDict(
                niveau=str(row.get("likert_niveau")) if row.get("likert_niveau") is not None else None,
                score=score_of(row.get("likert_niveau")),
                label=row.get("niveau_label"),
                situatieschets=row.get("situatieschets"),
                bron=row.get("bron_situatieschets"),
                toelichting=row.get("toelichting"),
            )
        )

    # Proxies are keyed on (effect, niveau) so they can hang under the matching
    # monetarisering niveau instead of forcing consumers to join two lists.
    proxies = defaultdict(list)
    for row in rows(wb[SHEETS["proxies"]]):
        niveau = str(row.get("likert_niveau")) if row.get("likert_niveau") is not None else None
        proxies[(row["effect_id"], niveau)].append(
            OrderedDict(
                stakeholder=row.get("stakeholder"),
                proxy=row.get("proxy"),
                bedrag=to_number(row.get("bedrag")),
                bedragTekst=row.get("bedrag"),
                eenheid=row.get("eenheid"),
                # Reeel of overdrachtsbetaling. Een overdracht is herverdeling
                # en hoort niet in een netto maatschappelijk totaal; zonder dit
                # veld is dat alleen uit de proxynaam af te leiden, en die zegt
                # het niet altijd (Bijstandsuitvoeringskosten is reeel) of in
                # wisselende bewoordingen. Leeg waar het workbook de kolom niet
                # heeft: onbekend, nooit "reeel".
                typePost=row.get("type_post"),
                bronBedrag=row.get("bron_bedrag"),
                bronEffectProxyRelatie=row.get("bron_effect_proxy_relatie"),
                toelichtingProxykeuze=row.get("toelichting_proxykeuze"),
                berekening=row.get("berekening"),
                aannames=row.get("aannames"),
                aannamescore=to_int(row.get("aannamescore")),
                overlapgroep=row.get("overlapgroep"),
            )
        )

    for row in rows(wb[SHEETS["monetarisering"]]):
        eid = row["effect_id"]
        niveau = str(row.get("likert_niveau")) if row.get("likert_niveau") is not None else None
        monetarisering = effecten[eid]["monetarisering"]
        monetarisering["document"] = row.get("monetarisering_document")
        monetarisering["eenheid"] = row.get("eenheid")
        monetarisering["niveaus"].append(
            OrderedDict(
                niveau=niveau,
                score=score_of(row.get("likert_niveau")),
                totaleWaardeIndicatief=to_number(row.get("totale_waarde_niveau_indicatief")),
                # De splitsing die bepaalt wat een consument mag optellen. Staat
                # hier en niet alleen op de proxyregels, zodat het antwoord op
                # "welk deel is maatschappelijk" geen inspectie van elke proxy
                # vergt.
                waarvanReeel=to_number(row.get("waarvan_reeel")),
                waarvanOverdracht=to_number(row.get("waarvan_overdracht")),
                berekening=row.get("berekening"),
                belangrijksteAannames=row.get("belangrijkste_aannames"),
                aannamescore=row.get("aannamescore"),
                proxies=proxies.pop((eid, niveau), []),
            )
        )

    if proxies:
        raise SystemExit(f"proxy rows without a matching monetarisering niveau: {sorted(proxies)}")

    return list(effecten.values())


def build_controle(effecten):
    """Cross-check each niveau total against the sum of its proxy amounts.

    The workbook derives `totale_waarde_niveau_indicatief` by summing the proxy
    rows, so a difference means the source itself is inconsistent (e.g. a
    percentage summed as if it were euros). We publish the discrepancies rather
    than silently reconciling them — this standard's whole point is traceability.
    """
    niet_gemonetariseerd, afwijkingen = [], []

    for effect in effecten:
        for niveau in effect["monetarisering"]["niveaus"]:
            tekstueel = [p for p in niveau["proxies"] if p["bedrag"] is None and p["bedragTekst"]]
            for proxy in tekstueel:
                niet_gemonetariseerd.append(
                    OrderedDict(
                        effectId=effect["id"],
                        niveau=niveau["niveau"],
                        proxy=proxy["proxy"],
                        bedragTekst=proxy["bedragTekst"],
                        eenheid=proxy["eenheid"],
                    )
                )

            som = round(sum(p["bedrag"] for p in niveau["proxies"] if p["bedrag"] is not None), 2)
            totaal = niveau["totaleWaardeIndicatief"]
            if totaal is not None and abs(som - totaal) > 0.01:
                afwijkingen.append(
                    OrderedDict(
                        effectId=effect["id"],
                        niveau=niveau["niveau"],
                        somProxybedragen=som,
                        totaleWaardeIndicatief=totaal,
                        verschil=round(totaal - som, 2),
                    )
                )

    return OrderedDict(
        toelichting=(
            "Automatische controles op deze versie. `nietGemonetariseerd` bevat proxyregels "
            "waarvan het bedrag geen enkelvoudig geldbedrag is (PM, n.v.t., een percentage); "
            "die tellen als 0 mee in de som. `somAfwijkingen` bevat niveaus waar de opgegeven "
            "`totaleWaardeIndicatief` afwijkt van de som van de proxybedragen — dat is een "
            "inconsistentie in de bron, niet in de conversie."
        ),
        nietGemonetariseerd=niet_gemonetariseerd,
        somAfwijkingen=afwijkingen,
    )


def build_bronnen(wb):
    bronnen = OrderedDict()
    for row in rows(wb[SHEETS["bronnen"]]):
        bid = row["bron_id"]
        bronnen[bid] = OrderedDict(
            id=bid,
            apaReferentie=row.get("apa_referentie"),
            bestand=row.get("bestand"),
            typeBron=row.get("type_bron"),
            relevantVoor=split_list(row.get("relevant_voor")),
            pagina=row.get("pagina_tabblad_tabel"),
            betrouwbaarheid=row.get("betrouwbaarheid"),
            opmerkingen=row.get("opmerkingen"),
        )
    return bronnen


def build_aggregatie(wb):
    """Sheet 8 holds the overlap table, then a worked example under its own header."""
    ws = wb[SHEETS["aggregatie"]]
    overlapgroepen, voorbeeld = [], OrderedDict(titel=None, regels=[], totaal=None, overlapcorrectie=None)
    section = "overlap"

    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = [clean(c) for c in row]
        if all(c is None for c in cells):
            continue
        first = cells[0]

        if isinstance(first, str) and first.startswith("VOORBEELD"):
            section, voorbeeld["titel"] = "voorbeeld", first
            continue
        if first == "overlapgroep":  # repeated header inside the example block
            continue

        if section == "overlap":
            overlapgroepen.append(
                OrderedDict(
                    overlapgroep=first,
                    cluster=cells[1],
                    aantalEffecten=to_int(cells[2]),
                    effecten=split_list(cells[3]),
                    alpha=to_number(cells[4]),
                    methode=cells[5],
                    correctieNodig=(cells[6] or "").lower() == "ja",
                )
            )
        elif first == "TOTAAL":
            voorbeeld["totaal"] = OrderedDict(naieveSom=to_number(cells[2]), gecorrigeerd=to_number(cells[5]))
        elif first == "Overlapcorrectie":
            voorbeeld["overlapcorrectie"] = to_number(cells[5])
        else:
            voorbeeld["regels"].append(
                OrderedDict(
                    overlapgroep=first,
                    aantalBijdragen=to_int(cells[1]),
                    naieveSom=to_number(cells[2]),
                    dominante=to_number(cells[3]),
                    alpha=to_number(cells[4]),
                    gecorrigeerd=to_number(cells[5]),
                )
            )

    return OrderedDict(overlapgroepen=overlapgroepen, voorbeeld=voorbeeld)


def build_parameters(wb):
    return [
        OrderedDict(
            id=slugify(row["parameter"]),
            parameter=row["parameter"],
            centraleWaarde=to_number(row.get("centrale_waarde")),
            centraleWaardeTekst=row.get("centrale_waarde"),
            peiljaar=row.get("peiljaar"),
            rangeLaag=to_number(row.get("range_laag")),
            rangeLaagTekst=row.get("range_laag"),
            rangeHoog=to_number(row.get("range_hoog")),
            rangeHoogTekst=row.get("range_hoog"),
            bron=row.get("bron"),
            gevoeligheid=(row.get("gevoeligheid") or "").lower() == "ja",
            opmerking=row.get("opmerking"),
        )
        for row in rows(wb[SHEETS["parameters"]])
    ]


def build_gevoeligheid(wb):
    """Sheet 10 ends with a free-text 'Interpretatie:' footer rather than a row."""
    scenarios, interpretatie = [], None
    for row in rows(wb[SHEETS["gevoeligheid"]]):
        driver = row.get("driver")
        if isinstance(driver, str) and driver.startswith("Interpretatie"):
            interpretatie = driver
            continue
        scenarios.append(
            OrderedDict(
                driver=driver,
                uitkomstMaat=row.get("uitkomst-maat"),
                laag=OrderedDict(waarde=row.get("laag_waarde"), uitkomst=to_number(row.get("laag_uitkomst"))),
                midden=OrderedDict(waarde=row.get("midden_waarde"), uitkomst=to_number(row.get("midden_uitkomst"))),
                hoog=OrderedDict(waarde=row.get("hoog_waarde"), uitkomst=to_number(row.get("hoog_uitkomst"))),
                toelichting=row.get("toelichting"),
            )
        )
    return OrderedDict(interpretatie=interpretatie, scenarios=scenarios)


def build_audit(wb):
    return [
        OrderedDict(
            onderdeel=row.get("onderdeel"),
            itemId=row.get("item_id"),
            probleem=row.get("probleem"),
            ernst=row.get("ernst"),
            voorgesteldeActie=row.get("voorgestelde_actie"),
            status=row.get("status"),
        )
        for row in rows(wb[SHEETS["audit"]])
    ]


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--sector", required=True)
    parser.add_argument("--version", required=True)
    parser.add_argument("--released-at", required=True, help="ISO date, e.g. 2026-08-17")
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    missing = [name for name in SHEETS.values() if name not in wb.sheetnames]
    if missing:
        raise SystemExit(f"workbook is missing required sheets: {missing}")

    effecten = build_effecten(wb, args.sector)

    document = OrderedDict(
        meta=OrderedDict(
            version=args.version,
            sector=args.sector,
            sectorLabel=SECTOR_LABELS.get(args.sector, args.sector.title()),
            releasedAt=args.released_at,
            source=f"Meetstandaard {SECTOR_LABELS.get(args.sector, args.sector)} (meetstandaard.xlsx)",
            generatedBy="tools/build-meetstandaard.py",
            toelichting=(
                "Release candidate voor versie 1.0. De inhoud is compleet en intern geauditeerd "
                "(zie `audit`), maar nog niet extern gevalideerd; onderdelen gemarkeerd als "
                "'needs verification' (nv) kunnen in 1.0 nog wijzigen. Bedragen zijn indicatieve "
                "maatschappelijke kosten (negatief) of baten (positief) per persoon per jaar ten "
                "opzichte van de nullijn. Tel proxybedragen niet zonder meer op: pas eerst de "
                "overlapcorrectie uit `aggregatie` toe."
            ),
        ),
        effecten=effecten,
        bronnen=build_bronnen(wb),
        parameters=build_parameters(wb),
        aggregatie=build_aggregatie(wb),
        gevoeligheid=build_gevoeligheid(wb),
        audit=build_audit(wb),
        controle=build_controle(effecten),
    )

    with open(args.out, "w", encoding="utf-8") as handle:
        json.dump(document, handle, ensure_ascii=False, indent=2)
        handle.write("\n")

    proxy_count = sum(len(n["proxies"]) for e in document["effecten"] for n in e["monetarisering"]["niveaus"])
    controle = document["controle"]
    print(
        f"Wrote {args.out}: {len(document['effecten'])} effecten, {proxy_count} proxyregels, "
        f"{len(document['bronnen'])} bronnen, {len(document['parameters'])} parameters"
    )
    for row in controle["somAfwijkingen"]:
        print(f"  warning: {row['effectId']} niveau {row['niveau']} total {row['totaleWaardeIndicatief']} != som {row['somProxybedragen']}")
    print(f"  {len(controle['nietGemonetariseerd'])} proxyregels zonder enkelvoudig geldbedrag")


if __name__ == "__main__":
    main()
