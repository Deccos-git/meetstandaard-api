#!/usr/bin/env python3
"""
Build a versioned interventiebibliotheek JSON document from the source workbook.

This is a different kind of standaard from the effect-based meetstandaarden: it
catalogues *interventions* and what they physically do, in three layers —

    Laag A  interventie -> fysiek effect   (kengetal per eenheid)
    Laag B  conversie                      (central prices/emission factors, tab Aannames)
    Laag C  monetarisatie                  (CO2 at the shadow price)

The workbook computes Laag B/C in Excel formulas, and those cells have **no
cached values** (the file has never been recalculated by Excel), so reading them
yields nothing. This script therefore recomputes them from the Aannames sheet
using the same formulas, and records both the inputs and the formula in the
output so a consumer can check the arithmetic rather than trust it.

Because the recomputation is only correct as long as it matches the workbook,
every formula this script re-implements is compared against the formula actually
in the cell. A workbook that starts multiplying by something else fails the
build instead of quietly producing different numbers.

Usage:
    python3 tools/build-interventiebibliotheek.py \
        --xlsx "/path/to/interventiebibliotheek.xlsx" \
        --standaard milieu-circulariteit \
        --label "Milieu & circulariteit" \
        --version 0.9 \
        --released-at 2026-08-17 \
        --out functions/data/interventiebibliotheek-milieu-circulariteit-0.9.json

Requires: openpyxl. Not a runtime dependency of the Cloud Functions — the
generated JSON is committed and imported directly.
"""

import argparse
import json
from collections import Counter, OrderedDict

import openpyxl

from xlsx_common import clean, numbered_rows, rounded, rows, slugify, to_number

# The four sheets that hold interventions. "Klimaat & Energie" is the only one
# with per-year physical factors; the rest carry a CO2 figure per unit.
INTERVENTIE_SHEETS = ["Klimaat & Energie", "Klimaat & Energie - overig", "Biodiversiteit & Natuur", "Circulariteit"]

# Aannames rows referenced by the workbook's formulas, by their B-column row.
AANNAME_ROWS = {
    "elektriciteitsprijs": 2,
    "gasprijs": 3,
    "waterprijs": 4,
    "emissiefactorElektriciteit": 5,
    "emissiefactorGas": 6,
    "emissiefactorWater": 7,
    "bestendigingsfactor": 8,
    "co2Schaduwprijs": 9,
    "groenestroomAandeel": 10,
    "meerprijsGroeneStroom": 11,
    "elektriciteitsprijsGewogen": 12,
}

# The one Aannames row the workbook derives from the others: households pay the
# grey price plus the green-certificate surcharge, weighted by how many of them
# have a green contract. It is the price the savings column uses — the emission
# factor deliberately stays location-based (workbook decision 2026-07-24).
ELEKTRICITEITSPRIJS_GEWOGEN_FORMULE = "=B2+B10*B11"

# The status-kengetal vocabulary, and the variants the workbook actually uses.
# Normalising here keeps consumers from having to know that "casusgebonden" and
# "casusgebonden / vergelijkend" are the same thing.
STATUS_ALIASES = {
    "casusgebonden": "casusgebonden / vergelijkend",
    "enabler-output": "enabler / output",
}

# Columns on the Klimaat & Energie sheet, by letter, so the formula check below
# can address the same cells the workbook does.
KE_KOLOM = {
    "elektra": "J",
    "besparingHuishouden": "K",
    "maatschappelijk": "L",
    "totaal": "M",
    "gas": "O",
    "water": "P",
    "bestendiging": "Q",
    "co2e": "R",
}


def verwacht_ke_formules(rij):
    """The four formulas the Klimaat & Energie sheet must hold on a given row.

    Written the way openpyxl reports them (no spaces), so they can be compared
    verbatim. Reproduced here rather than parsed because the point is to assert
    *which* aanname each term uses — that is exactly what a silent edit changes.
    """
    k = KE_KOLOM
    gas, elektra, water, bestendiging = k["gas"], k["elektra"], k["water"], k["bestendiging"]
    return {
        k["besparingHuishouden"]: (
            f"=({gas}{rij}*Aannames!$B$3+{elektra}{rij}*Aannames!$B$12+{water}{rij}*Aannames!$B$4)*{bestendiging}{rij}"
        ),
        k["co2e"]: (
            f"=({gas}{rij}*Aannames!$B$6+{elektra}{rij}*Aannames!$B$5+{water}{rij}*Aannames!$B$7)*{bestendiging}{rij}"
        ),
        k["maatschappelijk"]: f"={k['co2e']}{rij}*Aannames!$B$9",
        k["totaal"]: f"={k['besparingHuishouden']}{rij}+{k['maatschappelijk']}{rij}",
    }


def controleer_formules(ws, rij):
    """Fail the build when the workbook no longer computes what we recompute."""
    for kolom, verwacht in verwacht_ke_formules(rij).items():
        gevonden = ws[f"{kolom}{rij}"].value
        if gevonden != verwacht:
            raise SystemExit(
                f"Klimaat & Energie!{kolom}{rij}: workbook formula changed.\n"
                f"  verwacht: {verwacht}\n  gevonden: {gevonden}\n"
                "Update verwacht_ke_formules() and the recomputation together."
            )


def lees_bestendiging(waarde, aannames):
    """Bestendiging is either a literal or a reference to the central factor.

    The reference has no cached value, so reading it naively yields None and the
    old code silently fell back to 1 — publishing behaviour measures at their
    full first-year effect. Resolve it, or refuse to guess.
    """
    if waarde is None:
        return None
    if isinstance(waarde, str) and waarde.startswith("="):
        if waarde.replace(" ", "") != "=Aannames!$B$8":
            raise SystemExit(f"unknown bestendiging formula: {waarde}")
        return aannames["bestendigingsfactor"]
    return to_number(waarde)


def is_geverifieerd(bron):
    """The workbook marks its own confidence in the status text.

    "GEVERIFIEERDE bandbreedte ... midden = AANNAME" is not a verified figure:
    the range is sourced, the point estimate inside it is not. Same for anything
    still marked needs verification. Treating those as verified would overstate
    what this document knows.
    """
    tekst = str(bron or "")
    if not tekst.upper().startswith("GEVERIFIEERD"):
        return False
    return "AANNAME" not in tekst.upper() and "needs verification" not in tekst


def read_aannames(wb):
    """Central prices and emission factors — Laag B, and the formula inputs."""
    ws = wb["Aannames"]

    # The weighted electricity price is itself a formula in the workbook, so it
    # gets the same treatment as the interventie columns: verify, then recompute.
    gevonden = str(ws.cell(AANNAME_ROWS["elektriciteitsprijsGewogen"], 2).value or "").replace(" ", "")
    if gevonden != ELEKTRICITEITSPRIJS_GEWOGEN_FORMULE:
        raise SystemExit(
            f"Aannames!B{AANNAME_ROWS['elektriciteitsprijsGewogen']}: expected "
            f"{ELEKTRICITEITSPRIJS_GEWOGEN_FORMULE}, found {gevonden}"
        )

    waarden = {key: to_number(ws.cell(rij, 2).value) for key, rij in AANNAME_ROWS.items()}
    waarden["elektriciteitsprijsGewogen"] = rounded(
        waarden["elektriciteitsprijs"] + waarden["groenestroomAandeel"] * waarden["meerprijsGroeneStroom"], 6
    )

    afgeleid = {AANNAME_ROWS["elektriciteitsprijsGewogen"]: ELEKTRICITEITSPRIJS_GEWOGEN_FORMULE}

    lijst = []
    for nummer, row in numbered_rows(ws):
        formule = afgeleid.get(nummer)
        waarde = waarden["elektriciteitsprijsGewogen"] if formule else to_number(row.get("Waarde"))
        lijst.append(
            OrderedDict(
                id=slugify(row["Parameter"]),
                parameter=row["Parameter"],
                waarde=waarde,
                eenheid=row.get("Eenheid"),
                bron=row.get("Bron / status"),
                geverifieerd=is_geverifieerd(row.get("Bron / status")),
                # A derived parameter must show its derivation, not just a number.
                afgeleid=formule is not None,
                formule=formule,
            )
        )

    return lijst, waarden


def build_interventies(wb, a):
    """One record per intervention, with Laag B/C recomputed from Aannames."""
    interventies, seen = [], set()

    for sheet in INTERVENTIE_SHEETS:
        ws = wb[sheet]
        for nummer, row in numbered_rows(ws):
            naam = row.get("Interventie")
            if not naam:
                continue

            slug = slugify(naam)
            if slug in seen:
                raise SystemExit(f"duplicate interventie slug: {slug}")
            seen.add(slug)

            status = row.get("Status kengetal")
            gas = to_number(row.get("Gas (m3/jr)"))
            # The workbook labels this column "1. Energiebesparing (kWh/jr)",
            # but it is electricity specifically: it is multiplied by the
            # electricity price and the electricity emission factor, and it goes
            # negative (WG10) where a measure uses *more* power.
            elektra = to_number(row.get("1. Energiebesparing (kWh/jr)"))
            water = to_number(row.get("Water (m3/jr)"))
            bestendiging = lees_bestendiging(row.get("Bestendiging"), a)
            co2PerEenheid = to_number(row.get("CO2e (kg/eenheid)"))

            kengetallen = OrderedDict(
                gasM3PerJaar=gas,
                elektraKwhPerJaar=elektra,
                waterM3PerJaar=water,
                bestendiging=bestendiging,
                co2ePerEenheid=co2PerEenheid,
                co2ePerEenheidTekst=row.get("CO2e (kg/eenheid)"),
            )

            # The workbook's three outcomes, in its own order: what the household
            # saves, what society saves, and their sum.
            berekend = OrderedDict(
                besparingHuishoudenEurPerJaar=None,
                co2eKgPerJaar=None,
                maatschappelijkeBesparingEurPerJaar=None,
                totaleWaardeEurPerJaar=None,
                maatschappelijkeBesparingEurPerEenheid=None,
            )

            # Laag B: the Klimaat & Energie rows carry physical consumption, so
            # savings and emissions follow from the central prices and factors.
            if any(v is not None for v in (gas, elektra, water)):
                controleer_formules(ws, nummer)
                factor = 1 if bestendiging is None else bestendiging
                berekend["besparingHuishoudenEurPerJaar"] = rounded(
                    (
                        (gas or 0) * a["gasprijs"]
                        + (elektra or 0) * a["elektriciteitsprijsGewogen"]
                        + (water or 0) * a["waterprijs"]
                    )
                    * factor
                )
                berekend["co2eKgPerJaar"] = rounded(
                    (
                        (gas or 0) * a["emissiefactorGas"]
                        + (elektra or 0) * a["emissiefactorElektriciteit"]
                        + (water or 0) * a["emissiefactorWater"]
                    )
                    * factor
                )
                berekend["maatschappelijkeBesparingEurPerJaar"] = rounded(
                    berekend["co2eKgPerJaar"] * a["co2Schaduwprijs"]
                )
                berekend["totaleWaardeEurPerJaar"] = rounded(
                    berekend["besparingHuishoudenEurPerJaar"] + berekend["maatschappelijkeBesparingEurPerJaar"]
                )

            # Laag C: everything else is monetised straight off its CO2 figure.
            if co2PerEenheid is not None:
                berekend["maatschappelijkeBesparingEurPerEenheid"] = rounded(co2PerEenheid * a["co2Schaduwprijs"])

            interventies.append(
                OrderedDict(
                    id=row.get("Code") or slug,
                    slug=slug,
                    code=row.get("Code"),
                    domein=row.get("Domein"),
                    activiteitstype=row.get("Activiteitstype"),
                    interventie=naam,
                    eenheid=row.get("Eenheid"),
                    primaireEffecten=row.get("Primaire effecten"),
                    rekenmodel=row.get("Rekenmodel"),
                    monetarisatiebron=row.get("Monetarisatiebron"),
                    onderbouwing=row.get("Onderbouwing"),
                    bewijssterkte=row.get("Bewijssterkte"),
                    statusKengetal=STATUS_ALIASES.get(status, status),
                    statusKengetalBron=status,
                    afbakening=row.get("Afbakening / overlap"),
                    nietCo2Monetair=row.get("Niet-CO2 monetair (laag C)"),
                    wikipagina=row.get("Wikipagina"),
                    kengetallen=kengetallen,
                    berekend=berekend,
                )
            )

    return interventies


def build_berekening():
    """The formulas behind `berekend`, published so a consumer can redo them."""
    return OrderedDict(
        toelichting=(
            "Elke waarde in `berekend` is hier herrekend uit `aannames`, met dezelfde formule als in "
            "het werkboek. De namen tussen haakjes zijn `id`s uit `aannames`. Uitkomst 1 (energie in "
            "kWh) is geen berekening maar het kengetal zelf: `kengetallen.elektraKwhPerJaar`."
        ),
        besparingHuishoudenEurPerJaar=(
            "(gasM3PerJaar x gasprijs + elektraKwhPerJaar x elektriciteitsprijs-gewogen "
            "+ waterM3PerJaar x waterprijs) x bestendiging"
        ),
        co2eKgPerJaar=(
            "(gasM3PerJaar x emissiefactor-aardgas + elektraKwhPerJaar x emissiefactor-elektriciteit "
            "+ waterM3PerJaar x emissiefactor-drinkwater) x bestendiging"
        ),
        maatschappelijkeBesparingEurPerJaar="co2eKgPerJaar x co2-schaduwprijs",
        totaleWaardeEurPerJaar="besparingHuishoudenEurPerJaar + maatschappelijkeBesparingEurPerJaar",
        maatschappelijkeBesparingEurPerEenheid="co2ePerEenheid x co2-schaduwprijs",
        waarschuwing=(
            "totaleWaardeEurPerJaar telt een private besparing (lagere energierekening) op bij een "
            "maatschappelijke waarde (CO2 tegen de milieuprijs). Dat is een MKBA-totaal, geen kasgeld: "
            "rapporteer de twee componenten er altijd bij."
        ),
    )


def build_controle(interventies, aannames):
    """Publish what is not settled, rather than letting it look settled.

    Much of this library is explicitly provisional: the workbook's own rule is
    "geen kengetallen verzinnen — ontbreekt een bron, dan needs verification".
    Surfacing that count is the honest way to serve it.
    """
    zonder_kengetal = [
        OrderedDict(id=i["id"], interventie=i["interventie"], domein=i["domein"], tekst=i["kengetallen"]["co2ePerEenheidTekst"])
        for i in interventies
        if i["kengetallen"]["co2ePerEenheid"] is None
        and i["berekend"]["co2eKgPerJaar"] is None
    ]

    genormaliseerd = [
        OrderedDict(id=i["id"], van=i["statusKengetalBron"], naar=i["statusKengetal"])
        for i in interventies
        if i["statusKengetalBron"] != i["statusKengetal"]
    ]

    # Gedragsmaatregelen are published at their first-year effect times the
    # central persistence factor. Which rows that touches is a methodological
    # choice, so name them rather than leaving it in the arithmetic.
    met_bestendiging = [
        i["id"] for i in interventies if i["kengetallen"]["bestendiging"] not in (None, 1)
    ]

    return OrderedDict(
        toelichting=(
            "Automatische controles op deze versie. `zonderKengetal` bevat interventies waarvoor geen "
            "berekenbaar CO2-kengetal beschikbaar is (needs verification, casusgebonden of enabler); die "
            "leveren geen getal op en mogen niet als nul worden gelezen. `statusGenormaliseerd` toont "
            "statuswaarden die in de bron inconsistent geschreven waren en hier zijn gelijkgetrokken. "
            "`metBestendigingsfactor` bevat de gedragsmaatregelen waarop de centrale bestendigingsfactor "
            "is toegepast."
        ),
        aantalInterventies=len(interventies),
        zonderKengetal=zonder_kengetal,
        statusGenormaliseerd=genormaliseerd,
        metBestendigingsfactor=met_bestendiging,
        nietGeverifieerdeAannames=[a["parameter"] for a in aannames if not a["geverifieerd"]],
    )


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--standaard", required=True)
    parser.add_argument("--label", required=True)
    parser.add_argument("--version", required=True)
    parser.add_argument("--released-at", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    # Formulas, not cached values: the workbook has never been recalculated, so
    # data_only=True yields None for every computed cell. Reading the formula is
    # what makes the recomputation checkable — see controleer_formules().
    wb = openpyxl.load_workbook(args.xlsx, data_only=False)
    missing = [s for s in INTERVENTIE_SHEETS + ["Aannames", "Bronnen", "Afbakening"] if s not in wb.sheetnames]
    if missing:
        raise SystemExit(f"workbook is missing required sheets: {missing}")

    aannames, waarden = read_aannames(wb)
    interventies = build_interventies(wb, waarden)

    document = OrderedDict(
        meta=OrderedDict(
            version=args.version,
            standaard=args.standaard,
            label=args.label,
            # Distinguishes this from the effect-based meetstandaarden served by
            # the same endpoint: the document shape is entirely different.
            kind="interventiebibliotheek",
            releasedAt=args.released_at,
            source="Interventiebibliotheek maatschappelijke impact (interventiebibliotheek.xlsx)",
            generatedBy="tools/build-interventiebibliotheek.py",
            toelichting=(
                "Bibliotheek van interventies met hun fysieke effect per eenheid. Drie lagen: "
                "laag A het kengetal per interventie, laag B omrekening via de centrale prijzen en "
                "emissiefactoren in `aannames`, laag C monetarisering van CO2 tegen de milieuprijs. "
                "Per interventie staan drie uitkomsten in `berekend`: 1 energiebesparing (kWh, het "
                "kengetal zelf), 2 besparing voor het huishouden in euro, 3 maatschappelijke besparing "
                "in euro; `totaleWaardeEurPerJaar` is 2 + 3. Die kolommen zijn hier herberekend uit "
                "`aannames` volgens `berekening` — pas een prijs daar aan en alles schuift mee. Niet elk "
                "kengetal is hard: zie `bewijssterkte` en `statusKengetal` per interventie en het "
                "`controle`-blok. Tel binnen een overlapcluster niet op zonder de regels in `afbakening` "
                "toe te passen."
            ),
        ),
        domeinen=list(dict.fromkeys(i["domein"] for i in interventies if i["domein"])),
        interventies=interventies,
        aannames=aannames,
        berekening=build_berekening(),
        bronnen=[
            OrderedDict(
                id=slugify(row["Bron"]),
                bron=row["Bron"],
                domein=row.get("Domein"),
                laag=row.get("Laag"),
                toegang=row.get("Toegang"),
                wikipagina=row.get("Wikipagina"),
            )
            for row in rows(wb["Bronnen"])
        ],
        afbakening=[
            OrderedDict(onderwerp=row[0], toelichting=row[1])
            for row in ([clean(c) for c in r] for r in wb["Afbakening"].iter_rows(values_only=True))
            if row[0] or row[1]
        ],
        leeswijzer=[clean(r[0]) for r in wb["Leeswijzer"].iter_rows(values_only=True) if clean(r[0])],
        controle=build_controle(interventies, aannames),
    )

    with open(args.out, "w", encoding="utf-8") as handle:
        json.dump(document, handle, ensure_ascii=False, indent=2)
        handle.write("\n")

    per_domein = Counter(i["domein"] for i in interventies)
    berekend = sum(1 for i in interventies if i["berekend"]["co2eKgPerJaar"] is not None)
    print(f"Wrote {args.out}: {len(interventies)} interventies {dict(per_domein)}")
    print(f"  {berekend} met herberekende jaarcijfers, {len(document['controle']['zonderKengetal'])} zonder CO2-kengetal")
    print(f"  {len(document['bronnen'])} bronnen, {len(aannames)} aannames")


if __name__ == "__main__":
    main()
